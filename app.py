# <file name=1 path=/Users/ankurgupta/myprojects/mcc/app.py># app.py


import io, textwrap
import streamlit as st
import os
import pandas as pd
import matplotlib.pyplot as plt
from mcc import call_fracto, write_excel_from_ocr, _extract_rows, MAPPINGS, stamp_job_number

# ── Page config (must be first Streamlit command) ─────────────
st.set_page_config(
    page_title="PDF → Smart‑OCR → Excel",
    page_icon="📄",
    layout="wide",
)

# ── Fracto branding styles ────────────────────────────────────
FRACTO_PRIMARY   = "#0066FF"   # adjust if brand palette differs
FRACTO_DARK      = "#003B9C"
FRACTO_LIGHT_BG  = "#F5F8FF"

st.markdown(f"""
    <style>
    /* Page background */
    .stApp {{
        background: {FRACTO_LIGHT_BG};
    }}
    /* Center main content max-width 880px */
    .main .block-container{{
        max-width:880px;
        margin:auto;
    }}
    .block-container{{
        max-width:880px !important;
        margin-left:auto !important;
        margin-right:auto !important;
    }}
    /* Primary buttons */
    button[kind="primary"] {{
        background-color: {FRACTO_PRIMARY} !important;
        color: #fff !important;
        border: 0 !important;
    }}
    button[kind="primary"]:hover {{
        background-color: {FRACTO_DARK} !important;
        color: #fff !important;
    }}
    /* Header text color */
    h1 {{
        color: {FRACTO_DARK};
    }}
    /* Manual text_input boxes: white background & border */
    .stTextInput > div > div > input {{
        background-color: #ffffff !important;
        border: 1px solid #CCCCCC !important;
        border-radius: 4px !important;
    }}
    .stTextInput > div > div > input:focus {{
        border: 1px solid #0066FF !important;   /* Fracto primary on focus */
        box-shadow: 0 0 0 2px rgba(0,102,255,0.2) !important;
    }}
    /* File uploader box */
    .stFileUploader > div > div {{
        background-color: #ffffff !important;
        border: 1px solid #CCCCCC !important;
        border-radius: 4px !important;
        color: #222222 !important;
    }}
    /* Fix inside text in uploader */
    .stFileUploader label {{
        color: #222222 !important;
    }}
    /* Force background and text for all blocks */
    html, body, .stApp, .block-container {{
        background-color: #FFFFFF !important;
        color: #222222 !important;
    }}
    /* Buttons in login section */
    button, .stButton button {{
        background-color: #0066FF !important;
        color: #ffffff !important;
    }}
    button:hover, .stButton button:hover {{
        background-color: #003B9C !important;
        color: #ffffff !important;
    }}
    /* Labels stay dark text */
    label, .stMarkdown, .stSubheader, .stHeader, .stTextInput label {{
        color: #222222 !important;
    }}
    /* Password input */
    input[type="password"] {{
        background-color: #FFFFFF !important;
        color: #222222 !important;
        border: 1px solid #CCCCCC !important;
    }}
    /* Duplicate overrides in dark‑mode query */
    @media (prefers-color-scheme: dark) {{
        html, body, .stApp, .block-container {{
            background-color: #FFFFFF !important;
            color: #222222 !important;
        }}
        label, .stMarkdown, .stSubheader, .stHeader, .stTextInput label {{
            color: #222222 !important;
        }}
        input[type="password"] {{
            background-color: #FFFFFF !important;
            color: #222222 !important;
            border: 1px solid #CCCCCC !important;
        }}
    }}
    /* Force light theme when user is in dark mode */
    @media (prefers-color-scheme: dark) {{
        .stApp {{
            background: #FFFFFF !important;
        }}
        h1, h2, h3, h4, h5, h6, p, label, span, div, input, textarea {{
            color: #222222 !important;
        }}
        /* keep our primary buttons */
        button[kind="primary"] {{
            background-color: #0066FF !important;
            color: #fff !important;
        }}
        button[kind="primary"]:hover {{
            background-color: #003B9C !important;
        }}
        /* inputs */
        .stTextInput > div > div > input {{
            background-color: #ffffff !important;
            color: #222222 !important;
        }}
        /* uploader stays light */
        .stFileUploader > div > div {{
            background-color: #ffffff !important;
            border: 1px solid #CCCCCC !important;
            color: #222222 !important;
        }}
        .stFileUploader label {{
            color: #222222 !important;
        }}
    }}
    </style>
""", unsafe_allow_html=True)

st.markdown(
    """
    <style>
    .card-container {
        display: flex;
        gap: 1rem;
        flex-wrap: wrap;
        margin-bottom: 1rem;
    }
    .card {
        flex: 1 1 200px;
        background: #FFFFFF;
        border: 1px solid #E0E0E0;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        padding: 1rem;
        text-align: center;
    }
    .card-icon {
        font-size: 2rem;
        margin-bottom: 0.5rem;
    }
    .card h4 {
        margin: 0.2rem 0 0.5rem 0;
        color: #003B9C;
        font-weight: 600;
    }
    .card p {
        font-size: 0.9rem;
        line-height: 1.3rem;
        margin: 0;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# Logo banner at the top
st.image("fractologo.jpeg", width=180)

# ── Session keys ─────────────────────────────────────────────
if "excel_bytes" not in st.session_state:
    st.session_state["excel_bytes"] = None
if "excel_filename" not in st.session_state:
    st.session_state["excel_filename"] = ""
if "edited_excel_bytes" not in st.session_state:
    st.session_state["edited_excel_bytes"] = None
    st.session_state["edited_filename"] = ""

# ── Simple username/password gate ─────────────────────────────
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    st.subheader("🔐 Login required")
    uname = st.text_input("Username")
    pword = st.text_input("Password", type="password")
    if st.button("Login"):
        if uname == "mcc" and pword == "mcc@99":
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("Invalid credentials")
    st.stop()   # prevent the rest of the app from rendering

# Ensure FRACTO_API_KEY is available for mcc.call_fracto
if "FRACTO_API_KEY" in st.secrets:
    os.environ["FRACTO_API_KEY"] = st.secrets["FRACTO_API_KEY"]



# ── Hero / intro ──────────────────────────────────────────────
st.markdown(
    '''
    <div style="text-align:center;margin-bottom:32px;">
      <h2 style="color:#003B9C;font-weight:600;margin:0;">Automate imports. Eliminate re‑typing.</h2>
      <p style="font-size:1.05rem;line-height:1.5rem;margin:8px 0 24px;">
        Fracto converts your shipping invoices, customs docs and purchase orders into<br>
        ERP‑ready spreadsheets in seconds — complete with your business rules and validation checks.
      </p>
      <a href="#upload" style="
          background:#00AB6B;
          color:#fff;
          padding:10px 22px;
          border-radius:6px;
          text-decoration:none;
          font-weight:500;
          transition:background .2s;">
        Get started
      </a>
    </div>
    ''',
    unsafe_allow_html=True,
)

st.markdown("## Smart‑OCR to ERP‑ready Excel")

st.markdown('<h3 id="upload">1. Upload and process your PDF</h3>', unsafe_allow_html=True)

# ── Upload & Process ──────────────────────────────────────────
# Upload widget
pdf_file = st.file_uploader("Upload PDF", type=["pdf"])

# Manual fields (always visible)
st.markdown("#### Optional manual fields")
manual_inputs: dict[str, str] = {}
job_no: str | None = None

manual_fields = ["Job Number"]
for col in manual_fields:
    val = st.text_input(col, key=f"manual_{col}")
    if not val:
        continue
    if col == "Job Number":
        job_no = val          # only stamp on PDF
    else:
        manual_inputs[col] = val  # Excel overrides

# Process button
run = st.button("⚙️ Process PDF", disabled=pdf_file is None)

if run:
    if not pdf_file:
        st.warning("Please upload a PDF first.")
        st.stop()

    with st.spinner("Calling Fracto…"):
        pdf_bytes = pdf_file.read()
        if job_no:
            pdf_bytes = stamp_job_number(pdf_bytes, job_no)

        result = call_fracto(pdf_bytes, pdf_file.name)

        buffer = io.BytesIO()
        write_excel_from_ocr([result], buffer, overrides=manual_inputs)
        st.session_state["excel_bytes"]   = buffer.getvalue()
        st.session_state["excel_filename"] = pdf_file.name.replace(".pdf", "_ocr.xlsx")

    st.success("Excel generated!")

# ── Preview & download ────────────────────────────────────────
if st.session_state["excel_bytes"]:
    st.markdown("### 2. Review and export")
    st.download_button(
        "⬇️ Download original Excel",
        data=st.session_state["excel_bytes"],
        file_name=st.session_state["excel_filename"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_original",
    )

    df = pd.read_excel(io.BytesIO(st.session_state["excel_bytes"]))
    edited_df = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        key="editable_grid",
    )

    if st.button("💾 Save edits"):
        # Load original workbook to preserve formatting
        from openpyxl import load_workbook
        wb_orig = load_workbook(io.BytesIO(st.session_state["excel_bytes"]))
        ws      = wb_orig.active

        # Overwrite data rows (assumes header is row 1)
        for r_idx, (_, row) in enumerate(edited_df.iterrows(), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        out_buf = io.BytesIO()
        wb_orig.save(out_buf)
        st.session_state["edited_excel_bytes"] = out_buf.getvalue()
        st.session_state["edited_filename"] = st.session_state["excel_filename"].replace(
            ".xlsx", "_edited.xlsx"
        )
        st.success("Edits saved — scroll below to download the .xlsx file.")

    if st.session_state.get("edited_excel_bytes"):
        st.download_button(
            "⬇️ Download edited Excel",
            data=st.session_state["edited_excel_bytes"],
            file_name=st.session_state["edited_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_edited",
        )

    # ── Quick stats & visualisations ───────────────────────────
    view_df = edited_df if st.session_state.get("edited_excel_bytes") else df

    st.markdown("### 3. Quick stats")
    k1, k2 = st.columns(2)
    k1.metric("Total rows", view_df.shape[0])
    k2.metric("Blank cells", int(view_df.isna().sum().sum()))

    # Optionally show numeric totals if columns exist
    if "Qty" in view_df.columns:
        st.metric("Total Qty", f"{view_df['Qty'].sum():,.0f}")
    if "Unit Price" in view_df.columns:
        st.metric("Sum Unit Price", f"{view_df['Unit Price'].sum():,.0f}")

    # ── Top Part Numbers by Qty chart ─────────────────────────
    if {"Part No.", "Qty"}.issubset(view_df.columns):
        st.markdown("#### Top SKUs by Qty")
        top_qty = (
            view_df.groupby("Part No.")["Qty"]
            .sum(numeric_only=True)
            .sort_values(ascending=False)
            .head(10)
        )

        if top_qty.empty or top_qty.shape[0] < 1:
            st.info("No Qty data available to plot.")
        else:
            fig, ax = plt.subplots()
            top_qty.plot(kind="barh", ax=ax)
            ax.invert_yaxis()
            ax.set_xlabel("Qty")
            ax.set_ylabel("Part No.")
            st.pyplot(fig)

st.markdown("---")

# ── Benefits grid ─────────────────────────────────────────────
st.markdown("### Why choose **Fracto Imports**?")
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown("#### 🚀 10× Faster")
    st.write("Upload → processed Excel in under a minute, even for multi‑page PDFs.")
with col2:
    st.markdown("#### 🔍 Error‑free")
    st.write("AI‑assisted extraction + your manual overrides ensure 99.9 % accuracy.")
with col3:
    st.markdown("#### 🔗 Fits Your ERP")
    st.write("Column mapping matches your import template out‑of‑the‑box.")

st.markdown("---")


# ── Card rendering helper ─────────────────────────────────────
def render_card(icon: str, title: str, body: str, *, width="250px") -> str:
    """Return HTML for a single card."""
    return f"""
        <div class="card" style="max-width:{width};">
          <div class="card-icon">{icon}</div>
          <h4>{title}</h4>
          <p>{body}</p>
        </div>
    """

# ── How it works ──────────────────────────────────────────────
st.markdown('<h3 id="how">How it works</h3>', unsafe_allow_html=True)

steps = [
    ("📤", "Upload", "Drag PDFs or images of invoices, POs, customs docs into the drop‑zone."),
    ("🤖", "AI Extraction", "Vision models read tables, handwriting and stamps with 99 %+ accuracy."),
    ("📝", "Review & Edit", "Adjust any field inline — spreadsheet‑style editor keeps you in control."),
    ("🔄", "Export", "Download ERP‑ready Excel or push straight into your system via API."),
]

cols = st.columns(4)
for col, (icon, title, body) in zip(cols, steps):
    with col:
        col.markdown(render_card(icon, title, body), unsafe_allow_html=True)

st.markdown("---")

# ── Popular use‑cases ─────────────────────────────────────────
st.markdown('<h3 id="usecases">Popular use‑cases</h3>', unsafe_allow_html=True)

use_cases = [
    ("🛳️", "Import Logistics", "Bills of lading, packing lists, HS‑code mapping — ready for customs clearance."),
    ("🏭", "Manufacturing", "Supplier invoices and QC sheets flow directly into SAP/Oracle with serial‑level traceability."),
    ("💸", "Finance & AP", "Reconcile bank statements and purchase invoices 10× faster with zero manual key‑in."),
]

uc_cols = st.columns(3)
for col, (icon, title, body) in zip(uc_cols, use_cases):
    with col:
        col.markdown(render_card(icon, title, body, width="280px"), unsafe_allow_html=True)

st.markdown("---")

# ── Footer ────────────────────────────────────────────────────
st.markdown(
    "<div style='text-align:center;font-size:0.85rem;padding-top:2rem;color:#666;'>"
    "Made with ❤️ by <a href='https://www.fracto.tech' style='color:#0066FF;' target='_blank'>Fracto</a>"
    "</div>",
    unsafe_allow_html=True,
)