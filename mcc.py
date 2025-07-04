#!/usr/bin/env python
"""
fracto_page_ocr.py
──────────────────
Split a PDF page-by-page and pipe each page through Fracto Smart-OCR.
"""

import io
import os
import sys
import json
import time
import logging
from pathlib import Path
from typing import List, Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import yaml

import requests
from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas

# ─── PDF Stamping Helper ──────────────────────────────────────────────
def stamp_job_number(src_bytes: bytes, job_no: str, margin: int = 20) -> bytes:
    """
    Return new PDF bytes with an extra *margin* (pt) added to the top
    of every page, then stamps 'Job Number: <job_no>' inside that space.

    This ensures the stamp never covers the original page content.
    """
    if not job_no:
        return src_bytes

    from PyPDF2 import PdfReader, PdfWriter, Transformation, PageObject

    base_reader = PdfReader(io.BytesIO(src_bytes))
    writer      = PdfWriter()

    for orig_page in base_reader.pages:
        w = float(orig_page.mediabox.width)
        h = float(orig_page.mediabox.height)

        # 1️⃣  Create a new blank page taller by *margin*
        new_page = PageObject.create_blank_page(None, w, h + margin)

        # 2️⃣  Shift original page content down by `margin`
        orig_page.add_transformation(Transformation().translate(tx=0, ty=-margin))
        new_page.merge_page(orig_page)

        # 3️⃣  Create text overlay the same enlarged size
        overlay_buf = io.BytesIO()
        c = canvas.Canvas(overlay_buf, pagesize=(w, h + margin))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(40, h + margin - 15, f"Job Number: {job_no}")
        c.save()
        overlay_buf.seek(0)

        overlay_reader = PdfReader(overlay_buf)
        new_page.merge_page(overlay_reader.pages[0])

        # 4️⃣  Add to writer
        writer.add_page(new_page)

    out_buf = io.BytesIO()
    writer.write(out_buf)
    return out_buf.getvalue()

# ─── CONFIG ──────────────────────────────────────────────────────────────
FRACTO_ENDPOINT = "https://prod-ml.fracto.tech//upload-file-smart-ocr"
API_KEY         = os.getenv("FRACTO_API_KEY", "KUS-KUS-D09D77-709841-JXR4YETC")
PARSER_APP_ID   = "Tua4jrrYqLmCi7jt"
MODEL_ID        = "tv7"
EXTRA_ACCURACY  = "false"

# ──────────────────────────────────────────────────────────────────────────

logger = logging.getLogger("FractoPageOCR")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    datefmt="%H:%M:%S",
)

def _load_mapping():
    script_dir = Path(__file__).parent
    mapping_file = script_dir / "mapping.yaml"

    mappings = {}
    template_path = None
    sheet_name = None

    if mapping_file.exists():
        with open(mapping_file, "r", encoding="utf-8") as f:
            data = yaml.safe_load(f) or {}

        # Support both layouts:
        # 1. Wrapped under `excel_export`
        # 2. Flat mapping-only YAML
        if isinstance(data, dict) and "excel_export" in data:
            excel_cfg = data.get("excel_export", {})
            mappings       = excel_cfg.get("mappings", {})
            template_path  = excel_cfg.get("template_path")
            sheet_name     = excel_cfg.get("sheet_name")
        else:
            # Assume the YAML itself is the mapping dict
            mappings = data

        # Make template_path absolute if it exists
        if template_path:
            template_path = (script_dir / template_path).expanduser().resolve()

    return mappings, template_path, sheet_name

MAPPINGS, TEMPLATE_PATH, SHEET_NAME = _load_mapping()
HEADERS = list(MAPPINGS.keys())


def call_fracto(file_bytes: bytes, file_name: str) -> Dict[str, Any]:
    """
    Send the whole PDF to Fracto OCR and return the JSON response.
    """
    files = {
        "file": (file_name, io.BytesIO(file_bytes), "application/pdf"),
    }
    data = {
        "parserApp": PARSER_APP_ID,
        "model": MODEL_ID,
        "extra_accuracy": EXTRA_ACCURACY,
    }
    headers = {"x-api-key": API_KEY}

    try:
        start = time.time()
        resp = requests.post(
            FRACTO_ENDPOINT,
            headers=headers,
            files=files,
            data=data,
            timeout=600,         # seconds
        )
        resp.raise_for_status()
        elapsed = time.time() - start
        logger.info("✓ %s processed in %.2fs", file_name, elapsed)
        return {"file": file_name, "status": "ok", "data": resp.json()}
    except Exception as exc:
        logger.error("✗ %s failed: %s", file_name, exc)
        return {"file": file_name, "status": "error", "error": str(exc)}


def process_pdf(pdf_path: str) -> List[Dict[str, Any]]:
    """
    Send the complete PDF to Fracto and return a one‑element list with the API response.
    """
    pdf_path_obj = Path(pdf_path).expanduser().resolve()
    file_bytes = pdf_path_obj.read_bytes()
    result = call_fracto(file_bytes, pdf_path_obj.name)
    return [result]


# ─── Helper to persist results ───────────────────────────────────────────
def save_results(results: List[Dict[str, Any]], pdf_path: str, out_path: str | None = None) -> str:
    """
    Persist OCR results to disk.

    If *out_path* is None, a file named "<original‑stem>_ocr.json" is created
    alongside the input PDF.

    Returns the absolute path to the saved file.
    """
    if out_path is None:
        p = Path(pdf_path).expanduser().resolve()
        out_path = p.with_name(f"{p.stem}_ocr.json")
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(results, fh, indent=2)
    logger.info("Results written to %s", out_path)
    return str(out_path)


# ─── CLI ─────────────────────────────────────────────────────────────────
def _cli():
    """
    Usage:
        python -m mcc <pdf-path> [output.json] [output.xlsx] [KEY=VALUE ...]

    Convenience:
        • If you pass only two arguments and the second one ends with .xlsx / .xlsm / .xls,
          it is treated as the Excel output, and the JSON will default to
          "<pdf‑stem>_ocr.json" next to the PDF.
        • Any KEY=VALUE pairs will be written or overwritten in every row of the Excel output.
    """
    if len(sys.argv) < 2:
        print("Usage: python -m mcc <pdf-path> [output.json] [output.xlsx] [KEY=VALUE ...]")
        sys.exit(1)

    args = sys.argv[1:]

    pdf_path     = args[0]
    json_out     = None
    excel_out    = None

    # Collect KEY=VALUE overrides (e.g. --set Client=Acme or Client=Acme)
    overrides = {}
    remaining = []
    for arg in args[1:]:
        if "=" in arg:
            k, v = arg.split("=", 1)
            overrides[k.strip()] = v
        else:
            remaining.append(arg)
    # Re‑interpret remaining (non‑override) args for json/excel outputs
    if remaining:
        if remaining[0].lower().endswith((".xlsx", ".xlsm", ".xls")):
            excel_out = remaining[0]
        else:
            json_out = remaining[0]
    if len(remaining) >= 2:
        excel_out = remaining[1]

    if not os.path.isfile(pdf_path):
        logger.error("File not found: %s", pdf_path)
        sys.exit(2)

    results = process_pdf(pdf_path)

    # save JSON (use default if not supplied)
    save_results(results, pdf_path, json_out)

    # save Excel if requested
    if excel_out:
        write_excel_from_ocr(results, excel_out, overrides)




def _extract_rows(payload: Any) -> List[Dict[str, Any]]:
    """
    Heuristically extract a list[dict] rows from various JSON shapes
    that Fracto may return.

    • If *payload* is already a list of dicts → return as‑is.
    • If it's a dict, look for common keys ('data', 'rows', 'items', 'result', 'results')
      that hold a list of rows.
    • If the dict itself looks like a single row (has ≥1 HEADERS key) → wrap in list.
    Otherwise → empty list.
    """
    if isinstance(payload, list):
        return [r for r in payload if isinstance(r, dict)]

    if isinstance(payload, dict):
        for key in ("data", "rows", "items", "result", "results"):
            maybe = payload.get(key)
            if isinstance(maybe, list):
                return [r for r in maybe if isinstance(r, dict)]
        # treat dict itself as a single row if it shares keys
        if any(k in payload for k in HEADERS):
            return [payload]

    # Fallback: look inside `parsedData` for the first list of dicts
    if isinstance(payload, dict) and "parsedData" in payload:
        pd = payload["parsedData"]
        if isinstance(pd, dict):
            for v in pd.values():
                if isinstance(v, list) and v and isinstance(v[0], dict):
                    return [r for r in v if isinstance(r, dict)]

    return []  # fallback


def write_excel_from_ocr(
    results: List[Dict[str, Any]],
    output_path: str,
    overrides: dict[str, str] | None = None,
):
    """
    Write OCR rows to *output_path*.

    *overrides* – optional `{column_name: value}` pairs that will be:
      • Added to the header row if the column doesn't exist yet.
      • Written (or overwritten) with the given value in **every** row.
    """
    overrides = overrides or {}

    # Keep only overrides whose column exists in the YAML header list
    overrides = {k: v for k, v in overrides.items() if k in HEADERS}

    if TEMPLATE_PATH and TEMPLATE_PATH.exists():
        wb = load_workbook(TEMPLATE_PATH)
    else:
        wb = Workbook()
    if SHEET_NAME and SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        ws = wb.active

    # Clear existing data in the sheet
    ws.delete_rows(1, ws.max_row)

    # ── Gather all rows ──
    all_rows: list[dict] = []
    for result in results:
        payload = result.get("data", [])
        rows = _extract_rows(payload)
        all_rows.extend(rows)

    # Compose final header row (strictly from YAML)
    dynamic_headers = HEADERS  # no extras
    ws.append(dynamic_headers)

    # Write data
    written = 0
    for row in all_rows:
        values = []
        for col in dynamic_headers:
            field = MAPPINGS.get(col, col)  # default to same name if not mapped
            value = overrides.get(col, row.get(field, ""))
            values.append(value)
        ws.append(values)
        written += 1

    # ── Beautify ─────────────────────────────────────────────────────────
    # Header style: bold white text on dark blue
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="305496")  # MS Office blue
    header_align = Alignment(vertical="center", horizontal="center", wrap_text=True)

    # Thin border for all cells
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Apply header style
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Auto‑size columns based on max content length (capped)
    max_width = 60  # chars
    for col_idx, column in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        longest = max((len(str(cell.value)) if cell.value is not None else 0) for cell in column)
        width = min(max(longest + 2, 10), max_width)  # padding, min 10
        ws.column_dimensions[column[0].column_letter].width = width
        for cell in column[1:]:  # skip header, already styled
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Ensure gridlines visible (safe for ERP)
    ws.sheet_view.showGridLines = True

    # ── Beautify: borders, alignment, zebra striping ──
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Borders & alignment for all data cells (header + rows)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = center
    # Zebra striping: light grey fill on even data rows for readability
    stripe_fill = PatternFill("solid", fgColor="F2F2F2")  # very light grey
    for row_idx in range(2, ws.max_row + 1):  # data starts on row 2
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = stripe_fill
    # Freeze header row
    ws.freeze_panes = ws["A2"]
    wb.save(output_path)
    logger.info("Excel written to %s (%d rows, %d columns)", output_path, written, len(dynamic_headers))


# ─── Main Entry Point ────────────────────────────────────────────────────
if __name__ == "__main__":
    _cli()